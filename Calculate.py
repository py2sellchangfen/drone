# coding:utf-8
import xlrd
import sys
import untitled
from PyQt5.QtWidgets import QApplication, QDialog
import xlwt
from openpyxl import load_workbook, Workbook
import re
import numpy as np
import matplotlib.pyplot as plt
from PyQt5.QtGui import QPainter, QPixmap
from PyQt5.QtCore import Qt, QPoint
from matplotlib.backends.backend_qt5agg import FigureCanvasQTAgg as FigureCanvas
from matplotlib.figure import Figure
from matplotlib.animation import FuncAnimation
import pyqtgraph as pg

class MainDialog(QDialog):
    def __init__(self, file_name, sheet_name, parent=None):
        super(QDialog, self).__init__(parent)
        self.ui = untitled.Ui_Dialog()
        self.ui.setupUi(self)
        self.file_name = file_name
        self.sheet_name = sheet_name
        pg.setConfigOptions(leftButtonPan=False)
        pg.setConfigOption('background', 'w')
        pg.setConfigOption('foreground', 'k')
        a=0
    #读取界面样本设置信息（大方差or小方差）
    def GetSampleModel(self):
        spmodel = self.ui.SampleModelcomboBox.currentText()
        #print(spmodel)
        return spmodel

    def GetAlgorithmModel(self):
        amodel = self.ui.AlgorithmSelectcomboBox.currentText()
        return amodel

    def CreateTable(self, TableName, Sheetname1, sheetname2, Sheetname3):
        wb = Workbook()
        wb.create_sheet(Sheetname1, index=0)
        wb.create_sheet(sheetname2, index=1)
        wb.create_sheet(Sheetname3, index=2)
        wb.save(TableName)
        #work_book = xlwt.Workbook(encoding='utf-8')
        #work_book.save(TableName)


    def WriteSqmpleSheet(self, TableName, SheetName, Lable, SampleNum, DataSource1, DataSource2):
        #work_book = xlwt.Workbook(encoding='utf-8')
        #sheet = work_book.add_sheet(SheetName)
        wb = load_workbook(TableName)
        sheet = wb.active  #生成默认的工作簿
        sheet = wb[SheetName]
        sheet.
        houseids = []
        for id in range(SampleNum):
            houseids.append(id + 1)
        for i in range(SampleNum):
            #openpyxl调用excel时第一个单元格是（1，1）
            sheet.cell(row=i+1, column=1).value = houseids[i]
            sheet.cell(row=i+1, column=2).value = DataSource1[i]
            sheet.cell(row=i+1, column=3).value = DataSource2[i]
            sheet.cell(row=i+1, column=4).value = Lable

        wb.save(TableName)


    def ProductLable(self, TableName, SheetName, SampleNum, DataSource1, DataSource2):
        houselable = []
        wb = load_workbook(TableName)
        sheet = wb[SheetName]
        s = SampleNum-1

        for i in range(SampleNum):
            d1 = DataSource1[i]
            d2 = DataSource2[i]
            if d1>=140 and d2<=50:
                houselable.append(1)#高品质住房
            else:
                houselable.append(0)#中等品质住房
            sheet.cell(row=i+1, column=4).value = houselable[i]
        wb.save(TableName)


    def ProductSampleData(self):
        self.CreateTable('./ClassWork1Data.xlsx', 'SampleData1','SampleData2', 'TestData')
        space1 = [] #面积
        propertyfee1 = [] #物业费/百平米

        space2 = []  # 面积
        propertyfee2 = []  # 物业费/百平米

        spmodel = self.GetSampleModel()
        #print(spmodel)

        if spmodel=='小方差样本':
            space1 = np.random.normal(150, 2, 5000)
            propertyfee1 = np.random.normal(40, 2, 5000)

            space2 = np.random.normal(130, 2, 5000)
            propertyfee2 = np.random.normal(60, 2, 5000)

            samplenum = len(space1)



        else:
            space1 = np.random.normal(150, 20, 5000)
            propertyfee1 = np.random.normal(40, 20, 5000)

            space2 = np.random.normal(130, 20, 5000)
            propertyfee2 = np.random.normal(60, 20, 5000)

            samplenum = len(space1)
            # # 第0个工作簿的第0列写数据
            # self.WriteSqmpleSheet('ClassWork1Data.xlsx', 'SampleData1', 1, samplenum, space1, propertyfee1)  # 文件名、工作簿名、标签值、样本数、每一维内容
            # self.WriteSqmpleSheet('ClassWork1Data.xlsx', 'SampleData2', 0, samplenum, space2, propertyfee2)  # 文件名、工作簿名、工作簿序号、样本数、每一维内容
            # self.ui.textEdit.setPlainText(spmodel+'\n'+'样本生成成功！\n标签生成成功！')

        self.SourceSampleDatascatterPlot(samplenum, space1, propertyfee1, space2, propertyfee2)#画图
        # 第0个工作簿的第0列写数据
        self.WriteSqmpleSheet('ClassWork1Data.xlsx', 'SampleData1', 1, samplenum, space1, propertyfee1)  # 文件名、工作簿名、标签值、样本数、每一维内容
        # self.ProductLable('ClassWork1Data.xlsx', 'SampleData1',samplenum, space1, propertyfee1)
        self.WriteSqmpleSheet('ClassWork1Data.xlsx', 'SampleData2', 0, samplenum, space2, propertyfee2)  # 文件名、工作簿名、工作簿序号、样本数、每一维内容
        # self.ProductLable('ClassWork1Data.xlsx', 'SampleData2', samplenum, space2, propertyfee2)
        self.ui.textEdit.setPlainText(spmodel + '\n' + '样本生成成功！\n标签生成成功！')




    def SelectTrainData(self):
        #wb = load_workbook('ClassWork1Data.xlsx')
        #sheet_1 = wb['SampleData1']
        #sheet_2 = wb['SampleData2']
        wb: Workbook = load_workbook(self.file_name)
        sheet = wb[self.sheet_name]
        sheet = wb.active
        sheet.

        # # 打开Excel
        # workbook = xlrd.open_workbook('ClassWork1Data.xlsx')
        #
        # # 进入sheet
        # excel_sheet = workbook.sheet_by_index(0)
        #
        # # 获取行数和列叔
        # nrows_num = excel_sheet.nrows
        # ncols_num = excel_sheet.ncols
        # #生成随机行序号
        # random_row = np.random.randint(0, nrows_num, 2500, int)
        #
        # wb = load_workbook('ClassWork1Data.xlsx')
        # sheet1 = wb.get_sheet_by_name('SampleData1')
        # sheet_2 = wb['SampleData2']
        # wb.create_sheet('TrainData', index=2)
        # sheet_3 = wb['TrainData']
        # for i in random_row:
        #     sheet1.
        pass


    def TrainModel(self):
        amodel = self.GetAlgorithmModel()


        pass

    def PartitionSample(self):
        pass

    def PLAAlgorithm(self):
        pass

    def PocketAlgorithm(self):
        pass

    def SourceSampleDatascatterPlot(self, num, DataSource11, DataSource12, DataSource21, DataSource22):
        plt = pg.plot(title='散点图')
        x = []
        y = []
        for i in range(num):
            x.append(DataSource11[i])
            y.append(DataSource12[i])
        # plt = pg.PlotWidget(self)

        plt.plot(x, y, pen=None, symbol="x", symbolBrush='r')
        for i in range(num):
            x.append(DataSource21[i])
            y.append(DataSource22[i])
        plt.plot(x, y, pen=None, symbol="+", symbolBrush='b')

        pg.QtGui.QGuiApplication.exec_()
        # self.gridLayout.addWidget(plt, 1, 1, 1, 1)


if __name__ == '__main__':
    myapp = QApplication(sys.argv)
    myDlg = MainDialog()
    myDlg.show()

    sys.exit(myapp.exec_())